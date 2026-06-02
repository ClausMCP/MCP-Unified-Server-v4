#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
MCP Filesystem Advanced v1.1
Массовые операции (batch), универсальное извлечение текста (extract),
версионирование и глобальный FTS (Full-Text Search).
"""
import os
import shutil
import sqlite3
import hashlib
import time
from pathlib import Path
from datetime import datetime
from typing import List, Dict, Optional
from mcp_shared import (
    normalize_path, _ensure_allowed, BaseMCPServer, conversation_memory,
    dialog_ctx, _log, _format_size
)
from mcp_rate_limiter import safe_call

# ─── 4.1. Универсальный извлекатель текста ───────────────────────────────────
def extract_text_from_file(file_path: str, max_bytes: int = 100000) -> str:
    """Извлекает текст из файла в зависимости от его расширения."""
    p = Path(normalize_path(file_path))
    if not p.exists() or not p.is_file():
        return ""
    ext = p.suffix.lower()
    try:
        if ext == '.docx':
            try:
                from docx import Document
                doc = Document(str(p))
                return "\n".join(paragraph.text for paragraph in doc.paragraphs)
            except ImportError:
                return "[Error: python-docx not installed]"
        elif ext == '.xlsx':
            try:
                from openpyxl import load_workbook
                wb = load_workbook(str(p), read_only=True, data_only=True)
                text = []
                for sheet in wb.worksheets[:3]:
                    for row in sheet.iter_rows(values_only=True, max_row=1000):
                        text.append(" ".join(str(cell) for cell in row if cell is not None))
                wb.close()
                return "\n".join(text)
            except ImportError:
                return "[Error: openpyxl not installed]"
        elif ext == '.pdf':
            try:
                import PyPDF2
                text = []
                with open(p, 'rb') as f:
                    reader = PyPDF2.PdfReader(f)
                    for page in reader.pages[:10]:
                        text.append(page.extract_text() or "")
                return "\n".join(text)
            except ImportError:
                return "[Error: PyPDF2 not installed]"
            except Exception as e:
                return f"[Error reading PDF: {e}]"
        elif ext in ('.txt', '.py', '.md', '.json', '.csv', '.xml', '.log', '.ini', '.cfg'):
            with open(p, 'r', encoding='utf-8', errors='ignore') as f:
                return f.read(max_bytes)
        else:
            return ""
    except Exception as e:
        _log(f"Extract text error for {file_path}: {e}")
        return ""

def extract_text_from_folder(folder_path: str, extensions: List[str] = None, max_files: int = 50) -> Dict:
    """Извлекает текст из всех файлов в папке рекурсивно."""
    p = Path(normalize_path(folder_path))
    _ensure_allowed(p, "extract_text_from_folder")
    if not p.is_dir():
        return {"error": "Path is not a directory"}
    
    results = []
    count = 0
    exts = [e.lower() if e.startswith('.') else f'.{e.lower()}' for e in extensions] if extensions else None
    
    for f in p.rglob('*'):
        if count >= max_files:
            break
        if f.is_file():
            if exts and f.suffix.lower() not in exts:
                continue
            text = extract_text_from_file(str(f))
            if text:
                results.append({"file": str(f), "preview": text[:500]})
                count += 1
    return {"folder": str(p), "files_extracted": len(results), "results": results}

# ─── 4.2. Массовые операции с preview (dry-run) ──────────────────────────────
def batch_move_files(files: List[str] = None, target_dir: str = None, dry_run: bool = True, source_folder: str = None, pattern: str = "*") -> Dict:
    """Перемещает файлы, предварительно показывая что будет сделано."""
    if not files and source_folder:
        try:
            src_p = Path(normalize_path(source_folder))
            _ensure_allowed(src_p, "batch_move_files")
            files = [str(f) for f in src_p.rglob(pattern) if f.is_file()]
        except Exception as e:
            return {"status": "error", "message": f"Failed to find files: {e}"}
            
    if not files:
        return {"status": "error", "message": "No files to move."}
        
    target = Path(normalize_path(target_dir))
    _ensure_allowed(target, "batch_move_files")
    if not target.exists():
        if dry_run:
            return {"status": "dry_run", "message": f"Target directory will be created: {target}", "operations": []}
        target.mkdir(parents=True, exist_ok=True)
        
    operations = []
    errors = []
    for f in files:
        src = Path(normalize_path(f))
        if not src.exists():
            errors.append(f"Source not found: {f}")
            continue
        dst = target / src.name
        operations.append({"source": str(src), "destination": str(dst), "conflict": dst.exists()})
        
    if dry_run:
        return {"status": "dry_run", "operations": operations, "errors": errors}
        
    moved = []
    for op in operations:
        try:
            shutil.move(op["source"], op["destination"])
            moved.append(op)
        except Exception as e:
            errors.append(f"Failed to move {op['source']}: {e}")
            
    conversation_memory.add(
        op="batch_move_files", paths={"count": len(moved), "target": str(target)},
        status="moved", dialog=dialog_ctx.get(),
        context=f"Batch moved {len(moved)} files to {target}"
    )
    return {"status": "moved", "moved": moved, "errors": errors}

def batch_copy_files(files: List[str] = None, target_dir: str = None, dry_run: bool = True, source_folder: str = None, pattern: str = "*") -> Dict:
    """Копирует файлы, предварительно показывая что будет сделано."""
    if not files and source_folder:
        try:
            src_p = Path(normalize_path(source_folder))
            _ensure_allowed(src_p, "batch_copy_files")
            files = [str(f) for f in src_p.rglob(pattern) if f.is_file()]
        except Exception as e:
            return {"status": "error", "message": f"Failed to find files: {e}"}
            
    if not files:
        return {"status": "error", "message": "No files to copy."}
        
    target = Path(normalize_path(target_dir))
    _ensure_allowed(target, "batch_copy_files")
    if not target.exists():
        if dry_run:
            return {"status": "dry_run", "message": f"Target directory will be created: {target}", "operations": []}
        target.mkdir(parents=True, exist_ok=True)
        
    operations = []
    errors = []
    for f in files:
        src = Path(normalize_path(f))
        if not src.exists():
            errors.append(f"Source not found: {f}")
            continue
        dst = target / src.name
        operations.append({"source": str(src), "destination": str(dst), "conflict": dst.exists()})
        
    if dry_run:
        return {"status": "dry_run", "operations": operations, "errors": errors}
        
    copied = []
    for op in operations:
        try:
            src_path = Path(op["source"])
            if src_path.is_dir():
                shutil.copytree(src_path, op["destination"], dirs_exist_ok=True)
            else:
                shutil.copy2(src_path, op["destination"])
            copied.append(op)
        except Exception as e:
            errors.append(f"Failed to copy {op['source']}: {e}")
            
    conversation_memory.add(
        op="batch_copy_files", paths={"count": len(copied), "target": str(target)},
        status="copied", dialog=dialog_ctx.get(),
        context=f"Batch copied {len(copied)} files to {target}"
    )
    return {"status": "copied", "copied": copied, "errors": errors}

def batch_delete_files(files: List[str] = None, dry_run: bool = True, use_trash: bool = True, source_folder: str = None, pattern: str = "*") -> Dict:
    """Удаляет файлы, предварительно показывая что будет сделано."""
    if not files and source_folder:
        try:
            src_p = Path(normalize_path(source_folder))
            _ensure_allowed(src_p, "batch_delete_files")
            files = [str(f) for f in src_p.rglob(pattern) if f.is_file()]
        except Exception as e:
            return {"status": "error", "message": f"Failed to find files: {e}"}
            
    if not files:
        return {"status": "error", "message": "No files to delete."}
        
    operations = []
    errors = []
    for f in files:
        p = Path(normalize_path(f))
        if not p.exists():
            errors.append(f"Not found: {f}")
            continue
        operations.append({"path": str(p), "is_dir": p.is_dir(), "size": p.stat().st_size if p.is_file() else 0})
        
    if dry_run:
        return {"status": "dry_run", "operations": operations, "errors": errors, "use_trash": use_trash}
        
    deleted = []
    for op in operations:
        p = Path(op["path"])
        try:
            if use_trash:
                try:
                    from mcp_fs_trash import move_to_trash
                    move_to_trash(str(p), dialog_id=dialog_ctx.get())
                except ImportError:
                    if p.is_dir(): shutil.rmtree(p)
                    else: p.unlink()
            else:
                if p.is_dir(): shutil.rmtree(p)
                else: p.unlink()
            deleted.append(op)
        except Exception as e:
            errors.append(f"Failed to delete {op['path']}: {e}")
            
    conversation_memory.add(
        op="batch_delete_files", paths={"count": len(deleted)},
        status="deleted", dialog=dialog_ctx.get(),
        context=f"Batch deleted {len(deleted)} files"
    )
    return {"status": "deleted", "deleted": deleted, "errors": errors, "use_trash": use_trash}

# ─── 4.3. Версионирование (простое копирование с датой) ─────────────────────
def version_file(file_path: str, backup_dir: str = None) -> Dict:
    """Создаёт копию файла с суффиксом .YYYYMMDD_HHMMSS."""
    src = Path(normalize_path(file_path))
    _ensure_allowed(src, "version_file")
    if not src.exists() or not src.is_file():
        return {"status": "error", "message": f"File not found or not a file: {file_path}"}
    if not backup_dir:
        backup_dir = src.parent / ".versions"
    b_dir = Path(normalize_path(backup_dir))
    _ensure_allowed(b_dir, "version_file")
    b_dir.mkdir(parents=True, exist_ok=True)
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    dst = b_dir / f"{src.stem}.{timestamp}{src.suffix}"
    try:
        shutil.copy2(src, dst)
        conversation_memory.add(
            op="version_file", paths={"file": str(src), "version": str(dst)},
            status="versioned", dialog=dialog_ctx.get(),
            context=f"Created version {dst.name}"
        )
        return {"status": "versioned", "original": str(src), "version": str(dst), "size": dst.stat().st_size}
    except Exception as e:
        return {"status": "error", "message": f"Failed to create version: {e}"}

def list_versions(file_path: str, backup_dir: str = None) -> Dict:
    """Возвращает список всех версий файла."""
    src = Path(normalize_path(file_path))
    if not backup_dir:
        backup_dir = src.parent / ".versions"
    b_dir = Path(normalize_path(backup_dir))
    if not b_dir.exists():
        return {"status": "success", "versions": [], "count": 0}
    versions = []
    for f in b_dir.glob(f"{src.stem}.*{src.suffix}"):
        if f.is_file():
            versions.append({
                "path": str(f),
                "size": f.stat().st_size,
                "size_human": _format_size(f.stat().st_size),
                "modified": datetime.fromtimestamp(f.stat().st_mtime).isoformat()
            })
    versions.sort(key=lambda x: x["modified"], reverse=True)
    return {"status": "success", "versions": versions, "count": len(versions)}

# ─── 4.4. Улучшенная индексация содержимого (Глобальный FTS) ────────────────
def index_all_files_content(path: str, extensions: List[str] = None, force_reindex: bool = False, max_files: int = 500) -> Dict:
    """Индексирует содержимое всех поддерживаемых файлов для быстрого FTS."""
    service = "global_index"
    def _index():
        p = Path(normalize_path(path))
        _ensure_allowed(p, "index_all_files_content")
        if not p.is_dir():
            return {"error": "Path is not a directory"}
        index_db = p / ".global_index.db"
        conn = sqlite3.connect(str(index_db))
        conn.execute("""
            CREATE TABLE IF NOT EXISTS file_index (
                file_path TEXT PRIMARY KEY,
                content TEXT,
                hash TEXT,
                indexed_at REAL
            )
        """)
        conn.execute("CREATE VIRTUAL TABLE IF NOT EXISTS file_fts USING fts5(content, content=file_index)")
        exts = extensions or ['.txt', '.md', '.py', '.json', '.csv', '.xml', '.log', '.docx', '.xlsx', '.pdf']
        exts = [e.lower() if e.startswith('.') else f'.{e.lower()}' for e in exts]
        indexed = 0
        skipped = 0
        errors = 0
        count = 0
        for f in p.rglob('*'):
            if count >= max_files: break
            if not f.is_file(): continue
            ext = f.suffix.lower()
            if ext not in exts: continue
            count += 1
            try:
                current_hash = hashlib.md5(f"{f}_{f.stat().st_mtime}".encode()).hexdigest()
                row = conn.execute("SELECT hash FROM file_index WHERE file_path = ?", (str(f),)).fetchone()
                if not force_reindex and row and row[0] == current_hash:
                    skipped += 1
                    continue
                text = extract_text_from_file(str(f))
                if text:
                    conn.execute("INSERT OR REPLACE INTO file_index (file_path, content, hash, indexed_at) VALUES (?, ?, ?, ?)",
                                 (str(f), text[:100000], current_hash, time.time()))
                    indexed += 1
                else:
                    skipped += 1
            except Exception as e:
                _log(f"Global index error {f}: {e}")
                errors += 1
        conn.commit()
        conn.close()
        return {"indexed": indexed, "skipped": skipped, "errors": errors, "db_path": str(index_db)}
    return safe_call(service, _index)

def search_all_indexed_files(path: str, query: str, limit: int = 50) -> Dict:
    """Полнотекстовый поиск по проиндексированным файлам."""
    service = "global_search"
    def _search():
        p = Path(normalize_path(path))
        index_db = p / ".global_index.db"
        if not index_db.exists():
            return {"error": "Индекс не создан. Запустите index_all_files_content сначала."}
        conn = sqlite3.connect(str(index_db))
        try:
            cur = conn.execute("""
                SELECT file_path, content
                FROM file_fts
                JOIN file_index ON file_fts.rowid = file_index.rowid
                WHERE file_fts MATCH ?
                LIMIT ?
            """, (query, limit))
            results = [{"file": r[0], "snippet": r[1][:300]} for r in cur.fetchall()]
            return {"query": query, "results": results, "count": len(results)}
        finally:
            conn.close()
    return safe_call(service, _search)

# ─── Регистрация инструментов ───────────────────────────────────────────────
server = BaseMCPServer("filesystem-advanced", "1.1")
server.register_tool("extract_text_from_file", {
    "description": "Извлечь текст из файла (PDF, DOCX, XLSX, TXT, PY и др.)",
    "inputSchema": {"type": "object", "properties": {"file_path": {"type": "string"}, "max_bytes": {"type": "integer", "default": 100000}}, "required": ["file_path"]}
}, lambda **kw: extract_text_from_file(kw["file_path"], kw.get("max_bytes", 100000)))

server.register_tool("extract_text_from_folder", {
    "description": "Извлечь текст из всех файлов в папке рекурсивно",
    "inputSchema": {"type": "object", "properties": {"folder_path": {"type": "string"}, "extensions": {"type": "array", "items": {"type": "string"}}, "max_files": {"type": "integer", "default": 50}}, "required": ["folder_path"]}
}, lambda **kw: extract_text_from_folder(kw["folder_path"], kw.get("extensions"), kw.get("max_files", 50)))

server.register_tool("batch_move_files", {
    "description": "Массовое перемещение файлов с preview",
    "inputSchema": {"type": "object", "properties": {"files": {"type": "array", "items": {"type": "string"}}, "target_dir": {"type": "string"}, "dry_run": {"type": "boolean", "default": True}, "source_folder": {"type": "string"}, "pattern": {"type": "string", "default": "*"}}, "required": ["target_dir"]}
}, lambda **kw: batch_move_files(kw.get("files"), kw.get("target_dir"), kw.get("dry_run", True), kw.get("source_folder"), kw.get("pattern", "*")))

server.register_tool("batch_copy_files", {
    "description": "Массовое копирование файлов с preview",
    "inputSchema": {"type": "object", "properties": {"files": {"type": "array", "items": {"type": "string"}}, "target_dir": {"type": "string"}, "dry_run": {"type": "boolean", "default": True}, "source_folder": {"type": "string"}, "pattern": {"type": "string", "default": "*"}}, "required": ["target_dir"]}
}, lambda **kw: batch_copy_files(kw.get("files"), kw.get("target_dir"), kw.get("dry_run", True), kw.get("source_folder"), kw.get("pattern", "*")))

server.register_tool("batch_delete_files", {
    "description": "Массовое удаление файлов с preview",
    "inputSchema": {"type": "object", "properties": {"files": {"type": "array", "items": {"type": "string"}}, "dry_run": {"type": "boolean", "default": True}, "use_trash": {"type": "boolean", "default": True}, "source_folder": {"type": "string"}, "pattern": {"type": "string", "default": "*"}}}
}, lambda **kw: batch_delete_files(kw.get("files"), kw.get("dry_run", True), kw.get("use_trash", True), kw.get("source_folder"), kw.get("pattern", "*")))

server.register_tool("version_file", {
    "description": "Создать версию (бэкап) файла с суффиксом даты и времени",
    "inputSchema": {"type": "object", "properties": {"file_path": {"type": "string"}, "backup_dir": {"type": "string"}}, "required": ["file_path"]}
}, lambda **kw: version_file(kw["file_path"], kw.get("backup_dir")))

server.register_tool("list_versions", {
    "description": "Показать все сохранённые версии файла",
    "inputSchema": {"type": "object", "properties": {"file_path": {"type": "string"}, "backup_dir": {"type": "string"}}, "required": ["file_path"]}
}, lambda **kw: list_versions(kw["file_path"], kw.get("backup_dir")))

server.register_tool("index_all_files_content", {
    "description": "Индексировать содержимое всех файлов (PDF, Office, TXT) для FTS",
    "inputSchema": {"type": "object", "properties": {"path": {"type": "string"}, "extensions": {"type": "array", "items": {"type": "string"}}, "force_reindex": {"type": "boolean", "default": False}, "max_files": {"type": "integer", "default": 500}}, "required": ["path"]}
}, lambda **kw: index_all_files_content(kw["path"], kw.get("extensions"), kw.get("force_reindex", False), kw.get("max_files", 500)))

server.register_tool("search_all_indexed_files", {
    "description": "Быстрый полнотекстовый поиск по проиндексированным файлам",
    "inputSchema": {"type": "object", "properties": {"path": {"type": "string"}, "query": {"type": "string"}, "limit": {"type": "integer", "default": 50}}, "required": ["path", "query"]}
}, lambda **kw: search_all_indexed_files(kw["path"], kw["query"], kw.get("limit", 50)))

if __name__ == "__main__":
    server.run()